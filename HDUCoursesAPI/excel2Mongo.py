from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from course_model import Course
from new_parser import *
from db_mongo import DBMongo
from config import mongo_url

workbook = load_workbook('data/2021-2022.xlsx')
sheet: Worksheet = workbook['sheet1']

weekday_dict = {
    '1': '一',
    '2': '二',
    '3': '三',
    '4': '四',
    '5': '五',
    '6': '六',
    '7': '日'
}

courses = []
need_merge = {}

for one in sheet.iter_rows(min_row=2, max_row=5103, min_col=1, max_col=16):
    course = Course()
    teacher = one[6].value
    location = one[12].value
    other = one[14].value
    credit = float(one[9].value)
    key = teacher + '-' + location + '-' + other
    if ";" in location and credit >= 4.0:
        need_merge.setdefault(key, []).append(one)
        continue
    course.status = '已开'
    course.title = one[5].value
    course.credit = credit
    course.method = ''
    course.property = one[13].value
    course.teacher = teacher
    course.class_id = one[15].value
    time_info = one[3].value
    weekday = "周" + weekday_dict[one[2].value]
    location = parse_location(location)
    course.time_info = parse_time(time_info, weekday, location)
    week_info = one[4].value
    course.week_info = parse_week(week_info)
    course.location = [location]
    course.academic = one[11].value
    course.other = other
    courses.append(course.__dict__)

for key, item in need_merge.items():
    course = Course()
    if len(item) != 2:
        print(key)
        continue
    one = item[0]
    two = item[1]
    location = one[12].value.split(';')
    course.status = '已开'
    course.title = one[5].value
    course.credit = float(one[9].value)
    course.method = ''
    course.property = one[13].value
    course.teacher = one[6].value
    course.class_id = one[15].value
    time_info = one[3].value + "," + two[3].value
    weekday_one = "周" + weekday_dict[one[2].value]
    weekday_two = "周" + weekday_dict[two[2].value]
    weekday = [weekday_one, weekday_two]
    course.time_info = parse_dup_time(time_info, weekday, location)
    week_info = one[4].value
    course.week_info = parse_week(week_info)
    course.location = location
    course.academic = one[11].value
    course.other = one[14].value
    courses.append(course.__dict__)

db = DBMongo(mongo_url(), 'courses', 'coursetmp')
db.insert_many(courses)
